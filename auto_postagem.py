import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
import os
from dotenv import load_dotenv
import time
import pandas as pd
from tkinter import filedialog
import tkinter as tk
from openpyxl import load_workbook
import PyPDF2

load_dotenv()
url = os.getenv("URL")
usermane = os.getenv("USER")
Password = os.getenv("PASSWORD")

# Função para pedir o arquivo Excel
def pedir_arquivo():
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal do Tkinter

    # Abre a janela de seleção de arquivo
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione um arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    return caminho_arquivo

# Lê o arquivo Excel selecionado
caminho_arquivo = pedir_arquivo()
if caminho_arquivo:
    # Lê a planilha especificando o nome da aba
    df = pd.read_excel(caminho_arquivo, sheet_name='Planilha1')
else:
    print("Nenhum arquivo foi selecionado.")
    exit()

# Configurar opções do Chrome
chrome_options = Options()
chrome_options.add_experimental_option(
    "prefs",
    {
        "download.default_directory": os.path.join(os.getcwd(), "etiquetas_pdf"),
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
    }
)
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--headless")
# chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-extensions")

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

# driver.minimize_window()
time.sleep(5)
# Acessa o site
driver.get(url)

time.sleep(10)

# LOGIN
Usuario = driver.find_element(By.XPATH, '//*[@id="txtUsr"]')
Usuario.send_keys(usermane)

SenhaUser = driver.find_element(By.XPATH, '//*[@id="txtPwd"]')
SenhaUser.send_keys(Password)

BotaoLogin = driver.find_element(By.XPATH, '//*[@id="btnEfetuarLogin"]')
BotaoLogin.click()

# Abrindo a aba
time.sleep(2)
AbaPostagem = driver.find_element(By.XPATH, '/html/body/nav/div/ul/li[3]/a')
AbaPostagem.click()
BotaoPostagem = driver.find_element(By.XPATH, '/html/body/nav/div/ul/li[3]/ul/li[4]/a/span')
BotaoPostagem.click()

# Inicializa o índice
index = 0
total_linhas = len(df) + 2
i = 2

# Enquanto o índice for menor que o total de linhas
while index < total_linhas:
    
    # Atribui os valores de cada linha às variáveis
    cep = str(df.iloc[index]['Cep']).zfill(8)  # Preenche com zero à esquerda, se necessário
    # endereço = str(df.iloc[index]['Endereço'])
    # cidade = str(df.iloc[index]['Cidade'])
    # uf = str(df.iloc[index]['UF'])
    nome = str(df.iloc[index]['Nome'])
    numero = str(df.iloc[index]['Numero'])
    
    time.sleep(1.5)
    # Recarrega a página
    driver.refresh()
    time.sleep(0.5)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)
    botao_limpa = driver.find_element(By.XPATH, '//*[@id="cmdLimpar"]')
    botao_limpa.click()

    # Preenche os campos no site com os valores da linha
    print(cep)
    Cep_url = driver.find_element(By.XPATH, '//*[@id="txtCepDestinatario"]')
    Cep_url.send_keys(cep)

    Peso_url = driver.find_element(By.XPATH, '//*[@id="txtPeso"]')
    Peso_url.click()
    time.sleep(3)  # Aguarda o carregamento do campo
    Peso_url.send_keys("100")
    Peso_url.click()
    Peso_url.click()
    Cep_url.click()
    time.sleep(2)
    
    Embalagem = driver.find_element(By.XPATH, '//*[@id="frmEditarConhecimento"]/div/div/div/div[2]/div[3]/div[1]/div/div')
    print("Passou")
    time.sleep(1)
    Embalagem.click()
    time.sleep(2)
    Envelope = driver.find_element(By.XPATH, '//*[@id="frmEditarConhecimento"]/div/div/div/div[2]/div[3]/div[1]/div/div/div/ul/li[2]/a')
    Envelope.click()
    time.sleep(5)  # Aguarda a mudança de seleção

    Valor = driver.find_element(By.XPATH, '//*[@id="txtVlrDec"]')
    Valor.send_keys("26,50")  # Valor fixo, pode ser alterado conforme necessário

    NomeDoDistinaratario = driver.find_element(By.XPATH, '//*[@id="txtNomeDestinatario"]')
    NomeDoDistinaratario.click()
    time.sleep(4)
    NomeDoDistinaratario.send_keys(nome)
    

    NumEnderecoDoDistinaratario = driver.find_element(By.XPATH, '//*[@id="txtNumeroDestinatario"]')
    NumEnderecoDoDistinaratario.send_keys(numero)
    time.sleep(2)

    NumEnderecoDoDistinaratario.click
    time.sleep(2)
    botao_servico = driver.find_element(By.XPATH, '//*[@id="frmEditarConhecimento"]/div/div/div/div[2]/div[4]/div[4]/div/div[1]/button')
    time.sleep(2)
    botao_servico.click()
    time.sleep(2)
    botao_servico = None
    butao_servico1 = driver.find_element(By.XPATH, '//*[@id="frmEditarConhecimento"]/div/div/div/div[2]/div[4]/div[4]/div/div[1]/div/ul/li[1]/a/b[2]/span')
    butao_servico1.click()
    
    Botao_gravar = driver.find_element(By.XPATH, '//*[@id="frmEditarConhecimento"]/div/div/div/div[3]/div/div/button[1]')
    Botao_gravar.click()
    
    #Fora da automção web ///////////////////////////////////////////////////////////////////////////////////////////////////////////
    
    time.sleep(5)

    caminho_arquivo_excel = caminho_arquivo

# Define a pasta de destino como uma subpasta 'etiquetas' no diretório do script
    pasta_destino = os.path.join(os.getcwd(), "etiquetas_pdf")

# Cria a pasta caso ela não exista
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)

# Função para verificar o PDF mais recente na pasta
    def verificar_pdf_recente(pasta):
        arquivos_pdf = [f for f in os.listdir(pasta) if f.lower().endswith(".pdf")]
        if not arquivos_pdf:
            return None
    # Ordena os arquivos por data de modificação, do mais recente ao mais antigo
        arquivos_pdf.sort(key=lambda x: os.path.getmtime(os.path.join(pasta, x)), reverse=True)
        return arquivos_pdf[0]  # Retorna o PDF mais recente

# Função para processar o PDF e extrair os dados
    def processar_pdf(caminho_pdf):
        with open(caminho_pdf, "rb") as arquivo:
            leitor_pdf = PyPDF2.PdfReader(arquivo)
            texto = ""
            for pagina in leitor_pdf.pages:
                texto += pagina.extract_text()

    # Divide o texto extraído do PDF em linhas e obtém as informações necessárias
        linhas = texto.splitlines()
        codigo_rastreio = linhas[2]  # Ajuste de acordo com a linha correta
        nome_pdf = linhas[8]  # Ajuste de acordo com a linha correta
        return codigo_rastreio, nome_pdf

# Lê o arquivo Excel para atualizar
    df = pd.read_excel(caminho_arquivo_excel, sheet_name='Planilha1')

# Carrega o workbook para adicionar dados diretamente no Excel
    wb = load_workbook(caminho_arquivo_excel)
    ws = wb['Planilha1']

# Verifica e processa o PDF mais recente na pasta
    arquivo_pdf_recente = verificar_pdf_recente(pasta_destino)
    if arquivo_pdf_recente:
        caminho_arquivo_pdf = os.path.join(pasta_destino, arquivo_pdf_recente)

    # Processa o PDF e obtém as informações
        codigo_rastreio, nome_pdf = processar_pdf(caminho_arquivo_pdf)

    # Adiciona os dados extraídos ao DataFrame nas colunas E (codigo) e F (nome)
        ws.cell(row=i, column=4, value=str(codigo_rastreio))  # Coluna E
        ws.cell(row=i, column=5, value=str(nome_pdf))  # Coluna F

    # Adiciona o link para o PDF na coluna G
        link_pdf = f'=HYPERLINK("{caminho_arquivo_pdf}", "{arquivo_pdf_recente}")'
        ws.cell(row=i, column=6, value=link_pdf)  # Coluna G

    # Salva as alterações no arquivo Excel
        wb.save(caminho_arquivo_excel)
        print("Código de rastreio, nome do cliente e link do PDF mais recente adicionados com sucesso!")
        index += 1
        i += 1
        
    else:
        print("Nenhum PDF encontrado na pasta.")
        driver.quit()
        
